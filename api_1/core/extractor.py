from core.wrappers import safe_execute
from core.logger import info
import openpyxl
import PyPDF2


@safe_execute
def extract_text_from_xlsx(xlsx_path):
    """Extrai o texto de um arquivo XLSX e retorna como uma string."""
    info(f"Iniciando extração do arquivo XLSX: {xlsx_path}")
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    texto_total = ""

    for nome_planilha in workbook.sheetnames:
        planilha = workbook[nome_planilha]
        for linha in planilha.iter_rows(values_only=True):
            linha_texto = " ".join(
                [str(celula) for celula in linha if celula is not None]
            )
            texto_total += linha_texto + "\n"

    info(f"Extração do XLSX concluída: {xlsx_path}")
    return texto_total


@safe_execute
def extract_text_from_pdf(pdf_path):
    """Extrai o texto de um arquivo PDF e retorna como uma string."""
    info(f"Iniciando extração do arquivo PDF: {pdf_path}")
    with open(pdf_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        pdf_text = ""
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pdf_text += text + "\n"

    info(f"Extração do PDF concluída: {pdf_path}")
    return pdf_text


@safe_execute
def extract_content(file_path):
    """Extrai as transferências bancárias de um arquivo (XLSX ou PDF) e retorna como uma string."""
    info(f"Iniciando extração de conteúdo do arquivo: {file_path}")
    if file_path.endswith(".xlsx"):
        content = extract_text_from_xlsx(file_path)
    elif file_path.endswith(".pdf"):
        content = extract_text_from_pdf(file_path)
    else:
        raise ValueError("Formato de arquivo não suportado")
    info(f"Conteúdo extraído com sucesso do arquivo: {file_path}")
    return content
